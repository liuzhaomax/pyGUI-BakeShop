#!/usr/bin/python3
# coding=utf-8
'''''''''''''''''''''''''''''''''''''''''''''''''''
@FileName       Model.py
@Author         Zhao Liu
@StudId         30822750
@StartDate      06-09-2020
@LastModified   06-09-2020
@Function       Model base class
'''''''''''''''''''''''''''''''''''''''''''''''''''
from pandas import *
import openpyxl
from src.core.Object import *


class Model(Object):
    NAME = 'MODEL'

    def __init__(self, file_name, sheet_name):
        super().__init__()
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.table = None
        self.original_table = None
        self.new_data = None
        self.new_table_to_write = None

    @classmethod
    def get_model(cls, *args, **kwargs): # thread problem
        '''
        Get the instance of the class
        @param args: str, str
        @param kwargs:
        @return: Model or its children
        '''
        if not hasattr(cls, "_instance"):
            cls._instance = cls(*args, **kwargs)
        return cls._instance

    def read_excel(self):
        '''
        Read the excel file to form a dataframe
        @return:
        '''
        try:
            self.original_table = read_excel(self.file_name, sheet_name=self.sheet_name)
            self.table = self.original_table
        except:
            print('Excel reading failed.')

    def prepare_to_write(self, new_data):
        '''
        Abstract prepare_to_write
        @param new_data:
        @return:
        '''
        pass # assign new_table_to_write

    def write_excel(self, new_table_to_write):
        '''
        Write data to excel
        @param new_table_to_write: Dataframe
        @return:
        '''
        try:
            work_book = openpyxl.load_workbook(self.file_name)
            writer = ExcelWriter(self.file_name, engine='openpyxl')
            writer.book = work_book
            if self.sheet_name in writer.book.sheetnames:
                index = writer.book.sheetnames.index(self.sheet_name)
                writer.book.remove(writer.book.worksheets[index])
                writer.book.create_sheet(self.sheet_name, index)
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets} # copy existing sheets
            new_table_to_write.to_excel(writer, sheet_name=self.sheet_name, startrow=0, index=False, header=True)
            writer.save()
        except IOError as ex:
            print(ex)
            print("Write fails")
        finally:
            writer.close()